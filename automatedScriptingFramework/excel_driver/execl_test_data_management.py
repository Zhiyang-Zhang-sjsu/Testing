import openpyxl
from openpyxl.styles import PatternFill, Font

from web_keywords.web_UI_keywords import WebUIKeywords

TESTING_RESULT_ROW = 16
TESTING_RESULT_COL = 8

class ExeclTestDataManagement:
    def write_result(self, cell, row, col, status = False):
        if status:
            cell(row=row, column=col).value = "Passed"
            cell(row=row, column=col).font = Font(bold=True, color = "00FF00")
        else:
            cell(row=row, column=col).value = "Failed"
            cell(row=row, column=col).font = Font(bold=True, color = "FF0000")


    def run_test_data(self, file, log):
        driver = {}
        excel = openpyxl.load_workbook(file)
        try:
            for name in excel.sheetnames:
                sheet = excel[name]
                log.info(f"************{name}************")
                for values in sheet.values:
                    if type(values[0]) is int:
                        data = {}
                        data["by"] = values[2]
                        data["value"] = values[3]
                        data["text"] = values[4]
                        data['expected_result'] = values[6]

                        for key in list(data.keys()):
                            if data[key] is None:
                                del data[key]

                        log.info(f"Running: {values[5]}")
                        if values[1] == "open browser":
                            keywords = WebUIKeywords(data['text'], log)
                        else:
                            res = getattr(keywords, values[1])(**data)
                            if values[1] == "get_driver":
                                driver[name] = res
                            elif "assert" in values[1]:
                                self.write_result(sheet.cell, TESTING_RESULT_ROW, TESTING_RESULT_COL, res)
                                excel.save(file)
        except Exception as e:
            log.exception(f"Exception: {e}")
        finally:
            excel.close()

        return driver

